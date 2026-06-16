#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""创建 Luban 道具表示例（自动导入表 + 枚举）。"""

from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    raise SystemExit("请先安装: pip install openpyxl")

DATA_DIR = Path(__file__).resolve().parent / "Datas"


def write_schema_workbook(path: Path, sheet_name: str, headers_var: list, headers_type: list, data_rows: list | None = None) -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name
    for col, val in enumerate(headers_var, start=1):
        sheet.cell(row=1, column=col, value=val)
    for col, val in enumerate(headers_type, start=1):
        sheet.cell(row=2, column=col, value=val)
    if data_rows:
        for row_idx, row in enumerate(data_rows, start=3):
            for col_idx, val in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=val)
    wb.save(path)


def write_tables_xlsx(path: Path) -> None:
    write_schema_workbook(
        path,
        "table",
        ["##var", "full_name", "value_type", "index", "mode", "group", "comment", "read_schema_from_file", "input", "output", "tags"],
        ["##type", "string", "string", "string", "string", "string", "string", "bool", "string", "string", "string"],
    )


def write_beans_xlsx(path: Path) -> None:
    write_schema_workbook(
        path,
        "bean",
        ["##var", "full_name", "parent", "valueType", "sep", "alias", "comment", "tags", "group", "*fields", "", "", "", "", "", ""],
        ["##type", "string", "string", "bool", "string", "string", "string", "string", "string", "", "", "", "", "", ""],
    )


def write_enums_xlsx(path: Path) -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "enum"
    row1 = ["##var", "full_name", "flags", "unique", "group", "comment", "tags", "*items", "", "", ""]
    row2 = ["##var", "name", "alias", "value", "comment", "tags", "", "", "", "", ""]
    for col, val in enumerate(row1, start=1):
        sheet.cell(row=1, column=col, value=val)
    for col, val in enumerate(row2, start=1):
        sheet.cell(row=2, column=col, value=val)

    enum_rows = [
        ("", "EItemType", False, True, "c", "物品类型", "", "Weapon", "武器", 0, ""),
        ("", "", "", "", "", "", "", "Armor", "护甲", 1, ""),
        ("", "", "", "", "", "", "", "Consumable", "消耗品", 2, ""),
        ("", "", "", "", "", "", "", "Material", "材料", 3, ""),
        ("", "EQuality", False, True, "c", "品质", "", "White", "白", 0, ""),
        ("", "", "", "", "", "", "", "Green", "绿", 1, ""),
        ("", "", "", "", "", "", "", "Blue", "蓝", 2, ""),
        ("", "", "", "", "", "", "", "Purple", "紫", 3, ""),
    ]
    for row_idx, row in enumerate(enum_rows, start=3):
        for col_idx, val in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=val)
    wb.save(path)


def write_item_xlsx(path: Path) -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    fields = [
        ("id", "int", "道具ID"),
        ("name", "string", "名称"),
        ("desc", "string", "描述"),
        ("type", "EItemType", "类型"),
        ("quality", "EQuality", "品质"),
        ("max_stack", "int", "最大堆叠"),
        ("sell_price", "int", "出售价格"),
    ]
    sheet.cell(row=1, column=1, value="##var")
    sheet.cell(row=2, column=1, value="##type")
    sheet.cell(row=3, column=1, value="##")
    sheet.cell(row=4, column=1, value="##group")
    for i, (name, typ, comment) in enumerate(fields, start=2):
        sheet.cell(row=1, column=i, value=name)
        sheet.cell(row=2, column=i, value=typ)
        sheet.cell(row=3, column=i, value=comment)
        sheet.cell(row=4, column=i, value="c")

    data_rows = [
        (1001, "生命药水", "恢复少量生命值", "Consumable", "White", 99, 10),
        (1002, "铁剑", "普通铁制长剑", "Weapon", "White", 1, 100),
        (1003, "钢甲", "坚固钢制护甲", "Armor", "Green", 1, 250),
        (1004, "皮革", "普通皮革材料", "Material", "White", 999, 5),
    ]
    for row_idx, row in enumerate(data_rows, start=5):
        for col_idx, val in enumerate(row, start=2):
            sheet.cell(row=row_idx, column=col_idx, value=val)
    wb.save(path)


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    write_tables_xlsx(DATA_DIR / "__tables__.xlsx")
    write_beans_xlsx(DATA_DIR / "__beans__.xlsx")
    write_item_xlsx(DATA_DIR / "#Item-道具表.xlsx")
    enums_path = DATA_DIR / "__enums__.xlsx"
    if enums_path.exists():
        enums_path.unlink()
    print(f"已生成示例配表: {DATA_DIR}")


if __name__ == "__main__":
    main()
